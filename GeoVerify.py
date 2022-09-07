import streamlit as st
import requests
from urllib3.exceptions import InsecureRequestWarning
from io import BytesIO
import tempfile
from openpyxl import load_workbook
from openpyxl import Workbook
import re

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)
format = "JSON"
KEY = ""

global tokens
tokens = True

#Compares API results with data in spreadsheet. Adds new columns to track inconsistencies
def compareZiptoCity(zipColIdx, cityColIdx, addressCol, stateCol, numCols, ws):    
    # Create a new column. Do comparison and add result to same row and new column

    zipCodeMatchHeader = chr(65 + numCols) + "1"
    fullZipcodeHeader = chr(66 + numCols) + "1"
    cityMatchHeader = chr(67+ numCols) + "1"
    correctedCityHeader = chr(68+ numCols) + "1"
    finalCorrectHeader = chr(69+ numCols) + "1"

    ws[zipCodeMatchHeader] = "Zipcode Matches"
    ws[fullZipcodeHeader] = "Full Zipcode"
    ws[cityMatchHeader] = "City Matches"
    ws[correctedCityHeader] = "Corrected City"
    ws[finalCorrectHeader] = "V3 Corrected Address"
    
    for zipCode,cityName, address, state in zip(ws[zipColIdx], ws[cityColIdx], ws[addressCol], ws[stateCol]):
        if zipCode.row > 1 and cityName.row > 1:
            #Increment Column by 1 each iteration
            curCol = str(int(zipCodeMatchHeader[-1]) + 1)
            zipCodeMatchHeader = re.sub(r".$", curCol, zipCodeMatchHeader)
            fullZipcodeHeader = re.sub(r".$", curCol, fullZipcodeHeader)
            cityMatchHeader = re.sub(r".$", curCol, cityMatchHeader)
            correctedCityHeader = re.sub(r".$", curCol, correctedCityHeader)
            finalCorrectHeader = re.sub(r".$", curCol, finalCorrectHeader)
            URL = f"""https://geodata.cdxtech.com/api/geoverify?key={KEY}&address={address.value}&citystatezip={cityName.value}-{state.value}-{zipCode.value}&format={format}"""
            res = requests.get(URL, verify=False)
            apiCity = None
            
            if res:
                res = res.json()
            else:
                res = None

            #We found a match

            if res and not res["results"]:
                st.error("Out of tokens!")
                global tokens
                tokens = False
                break
            if res and res["results"][0]["errorMessage"] == None:
                apiCity = res["results"][0]["preferredCity"].lower()

                #Zipcode Matches?
                if str(res["results"][0]["zipcode"]).strip() == str(zipCode.value).strip() or str(res["results"][0]["nineDigitZipcode"]).strip() == str(zipCode.value).strip():
                    ws[zipCodeMatchHeader] = "True"
                else:
                    ws[zipCodeMatchHeader] = "False"
                ws[fullZipcodeHeader] = res["results"][0]["nineDigitZipcode"]
                #USPS City does not match ZipCode
                if apiCity == res["results"][0]["city"].lower():
                    ws[cityMatchHeader] = "True"
                else:
                    ws[cityMatchHeader] = "False"
                    ws[correctedCityHeader] = res["results"][0]["preferredCity"]
                
                tempAddressStr = res["results"][0]["fullAddressOut"].split(",")
                finalAddressStr = f"""{tempAddressStr[0]}, {res["results"][0]["preferredCity"]} {tempAddressStr[2]}"""
                ws[finalCorrectHeader] = finalAddressStr
            else:
                ws[zipCodeMatchHeader] = "Unable to determine address."
            

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        workbook.save(tmp.name)
        data = BytesIO(tmp.read())
        tmp.flush()
        tmp.close()
    return data
                


#GUI starts here
st.title("Validate Addresses")
uploaded_file = st.file_uploader("Upload .xlsx file here", type="xlsx", accept_multiple_files=False, key=None, help=None, on_change=None, args=None, kwargs=None, disabled=False)

KEY = st.text_input("Enter CDX API Key")

if uploaded_file:
    
    workbook = load_workbook(uploaded_file, data_only=True)
    selectedSheet = st.selectbox("Select sheet", workbook.sheetnames, 0)
    ws = workbook[selectedSheet]
    
    #ws = workbook.active
    numCols = ws.max_column
    numRows = ws.max_row

    #Key is name, value is Column index
    headers = {}
    currIdx = 65
    for cell in ws[1]:
        headers[cell.value] = chr(currIdx)
        currIdx += 1


    
    zipCol = st.selectbox("Zip Code Column", headers.keys(), 1)
    zipColIdx = headers[zipCol]
    st.write("Selected Zip Code Column: ", zipCol)

    cityCol = st.selectbox("City Column", headers.keys(), 1)
    cityColIdx = headers[cityCol]
    st.write("Selected Zip Code Column: ", cityCol)

    addressCol = st.selectbox("Address Column", headers.keys(), 1)
    addressColIdx = headers[addressCol]
    st.write("Selected Address Column: ", addressCol)

    stateCol = st.selectbox("State Column", headers.keys(), 1)
    stateColIdx = headers[stateCol]
    st.write("Selected State Column: ", stateCol)

    data = None
    clicked = st.button("Confirm Choices", help="Click this to confirm choices")
    if clicked:
        with st.spinner("Running API calls..."):
            data = compareZiptoCity(zipColIdx, cityColIdx,addressColIdx,stateColIdx, numCols, ws)
    if data and tokens:
        st.download_button(
            label="Download Excel Workbook",
            data = data,
            file_name="workbook.xlsx",
            mime="xlsx"
        )

